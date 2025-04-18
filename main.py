import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

ARQUIVO_PLANILHA = "produtos.xlsx"

# ------------------ UTILIDADES ------------------

def carregar_planilha():
    if not os.path.exists(ARQUIVO_PLANILHA):
        planilha = Workbook()
        folha = planilha.active
        folha.append(["Nome", "Quantidade", "Validade"])
        planilha.save(ARQUIVO_PLANILHA)
    return load_workbook(ARQUIVO_PLANILHA)

def salvar_planilha(planilha):
    planilha.save(ARQUIVO_PLANILHA)

def ordenar_por_validade(produtos):
    def validade_para_data(validade):
        try:
            return datetime.strptime(validade, "%m/%Y")
        except:
            return datetime.max
    return sorted(produtos, key=lambda x: validade_para_data(x[2]))

def atualizar_visualizacao_validade(tree):
    for item in tree.get_children():
        tree.delete(item)

    planilha = carregar_planilha()
    folha = planilha.active
    produtos = [row for row in folha.iter_rows(min_row=2, values_only=True) if row[1] > 0]
    produtos_ordenados = ordenar_por_validade(produtos)

    for nome, quantidade, validade in produtos_ordenados:
        tree.insert("", "end", values=(nome, quantidade, validade))

# ------------------ INTERFACE PRINCIPAL ------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("📦 Gerenciador de Produtos")
        self.geometry("550x550")
        self.configure(bg="#f7f7f7")

        self.frames = {}

        for Tela in (TelaInicial, TelaAdicionar, TelaRemover, TelaCadastrar):
            frame = Tela(self)
            self.frames[Tela] = frame
            frame.place(x=0, y=0, relwidth=1, relheight=1)

        self.mostrar_tela(TelaInicial)

    def mostrar_tela(self, tela):
        frame = self.frames[tela]
        frame.tkraise()
        if tela == TelaInicial:
            frame.atualizar()

# ------------------ TELA INICIAL ------------------

class TelaInicial(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")

        tk.Label(self, text="📦 Gerenciador de Produtos", font=("Segoe UI", 16, "bold"), bg="#f7f7f7").pack(pady=20)

        botoes = tk.Frame(self, bg="#f7f7f7")
        botoes.pack(pady=10)

        tk.Button(botoes, text="➕ Adicionar Itens", width=20, command=lambda: master.mostrar_tela(TelaAdicionar)).grid(row=0, column=0, padx=10)
        tk.Button(botoes, text="➖ Remover Itens", width=20, command=lambda: master.mostrar_tela(TelaRemover)).grid(row=0, column=1, padx=10)
        tk.Button(botoes, text="📋 Cadastrar Itens", width=20, command=lambda: master.mostrar_tela(TelaCadastrar)).grid(row=0, column=2, padx=10)

        tk.Label(self, text="🗓 Produtos por validade:", font=("Segoe UI", 12), bg="#f7f7f7").pack(pady=(30, 10))

        self.tree = ttk.Treeview(self, columns=("Nome", "Quantidade", "Validade"), show="headings", height=12)
        for col in ("Nome", "Quantidade", "Validade"):
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=150)
        self.tree.pack()

    def atualizar(self):
        atualizar_visualizacao_validade(self.tree)

# ------------------ TELA ADICIONAR ------------------

class TelaAdicionar(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.master = master

        tk.Label(self, text="➕ Adicionar Produto", font=("Segoe UI", 14, "bold"), bg="#f7f7f7").pack(pady=20)

        self.nome = self.criar_entrada("Nome")
        self.quantidade = self.criar_entrada("Quantidade")
        self.validade = self.criar_entrada("Validade (MM/AAAA)")

        tk.Button(self, text="Salvar", width=20, bg="#4CAF50", fg="white", command=self.salvar).pack(pady=10)
        tk.Button(self, text="← Voltar", command=lambda: self.master.mostrar_tela(TelaInicial)).pack()

    def criar_entrada(self, texto):
        tk.Label(self, text=texto, font=("Segoe UI", 10), bg="#f7f7f7").pack()
        entrada = tk.Entry(self, font=("Segoe UI", 10), width=40)
        entrada.pack(pady=5)
        return entrada

    def salvar(self):
        nome = self.nome.get().strip()
        validade = self.validade.get().strip()
        try:
            quantidade = int(self.quantidade.get())
        except:
            messagebox.showerror("Erro", "Quantidade inválida.")
            return

        planilha = carregar_planilha()
        folha = planilha.active
        atualizado = False
        for row in folha.iter_rows(min_row=2):
            if row[0].value.strip().lower() == nome.lower() and row[2].value == validade:
                row[1].value += quantidade
                atualizado = True
                break
        if not atualizado:
            folha.append([nome, quantidade, validade])

        salvar_planilha(planilha)
        self.nome.delete(0, tk.END)
        self.quantidade.delete(0, tk.END)
        self.validade.delete(0, tk.END)
        messagebox.showinfo("Salvo", "Produto adicionado.")
        self.master.mostrar_tela(TelaInicial)

# ------------------ TELA REMOVER ------------------

class TelaRemover(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")
        self.master = master

        tk.Label(self, text="➖ Remover Produto", font=("Segoe UI", 14, "bold"), bg="#f7f7f7").pack(pady=20)

        self.nome = self.criar_entrada("Nome")
        self.quantidade = self.criar_entrada("Quantidade a Remover")

        tk.Button(self, text="Remover", width=20, bg="#f44336", fg="white", command=self.remover).pack(pady=10)
        tk.Button(self, text="← Voltar", command=lambda: self.master.mostrar_tela(TelaInicial)).pack()

    def criar_entrada(self, texto):
        tk.Label(self, text=texto, font=("Segoe UI", 10), bg="#f7f7f7").pack()
        entrada = tk.Entry(self, font=("Segoe UI", 10), width=40)
        entrada.pack(pady=5)
        return entrada

    def remover(self):
        nome = self.nome.get().strip()
        try:
            quantidade = int(self.quantidade.get())
            if quantidade <= 0:
                raise ValueError
        except:
            messagebox.showerror("Erro", "Quantidade inválida.")
            return

        planilha = carregar_planilha()
        folha = planilha.active
        restante = quantidade
        alterado = False

        for row in folha.iter_rows(min_row=2):
            if row[0].value.strip().lower() == nome.lower():
                qnt = row[1].value
                if qnt >= restante:
                    row[1].value = qnt - restante
                    alterado = True
                    break
                else:
                    restante -= qnt
                    row[1].value = 0
                    alterado = True

        if alterado:
            salvar_planilha(planilha)
            messagebox.showinfo("Sucesso", f"{quantidade} unidade(s) removida(s).")
        else:
            messagebox.showwarning("Não encontrado", "Produto não encontrado ou sem estoque.")

        self.nome.delete(0, tk.END)
        self.quantidade.delete(0, tk.END)
        self.master.mostrar_tela(TelaInicial)

# ------------------ TELA CADASTRAR (EXTRAS FUTUROS) ------------------

class TelaCadastrar(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="#f7f7f7")

        tk.Label(self, text="📋 Cadastro de Produtos", font=("Segoe UI", 14, "bold"), bg="#f7f7f7").pack(pady=20)
        tk.Label(self, text="Funcionalidade futura para cadastro detalhado", font=("Segoe UI", 10), bg="#f7f7f7").pack(pady=10)
        tk.Button(self, text="← Voltar", command=lambda: master.mostrar_tela(TelaInicial)).pack(pady=20)

# ------------------ EXECUÇÃO ------------------

if __name__ == "__main__":
    app = App()
    app.mainloop()
